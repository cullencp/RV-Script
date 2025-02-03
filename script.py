import openpyxl
from openpyxl.styles import Alignment, Font
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import os
import sys
from tkinter import PhotoImage
from openpyxl.drawing.image import Image as XLImage  # For inserting images

# Constants for cell references and formatting
HEADER_KEYWORDS = [
    "tag", "manufacturer", "model", "process connection", "immersion length",
    "control signal", "min range", "max range", "unit", "order code",
    "valve make / model number", "actuator make / model number",
    "line size", "dial setting", "flow rate"
]
DEFAULT_START_ROW = 6
FONT_SIZE = 11

# --- Resource Path Helper ---
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller."""
    try:
        base_path = sys._MEIPASS  # PyInstaller temporary folder
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- Unique Identifier Lookup Helper ---
def get_instrument_no_index(column_indices):
    """Return the index for the unique identifier column using several possible keys.
       This function does not use the 'or' chain so that an index of 0 is accepted."""
    possible_keys = ["no.", "no", "item no.", "item no", "no:", "item no:"]
    for key in possible_keys:
        if key in column_indices:
            return column_indices[key]
    return None

# --- Excel and Template Functions ---
def get_sheet_by_partial_name(wb, partial_name):
    """Find a sheet by partial name match."""
    for sheet in wb.sheetnames:
        if partial_name.lower() in sheet.lower():
            return sheet
    raise ValueError(f"Sheet with partial name '{partial_name}' not found.")

def format_value(value):
    """Format the value in uppercase or return 'N/A' if empty."""
    if value is None or (isinstance(value, str) and value.strip().lower() == "n/a"):
        return "N/A"
    if isinstance(value, str):
        return value.upper()
    return value

def detect_header_row(sheet):
    """Dynamically detect the header row based on keyword matching."""
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if any(any(keyword in str(cell).lower() for keyword in HEADER_KEYWORDS) for cell in row):
            return row_idx
    raise ValueError("Header row not found in the first 10 rows.")

def get_column_indices(sheet, header_row):
    """Get a dictionary mapping header names to column indices."""
    header_row_data = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    print("Detected Header Row:", header_row_data)  # Debug statement
    column_indices = {}
    for idx, header in enumerate(header_row_data):
        if header:
            # Print each header with its index for debugging purposes.
            print(f"Header at index {idx}: {repr(header)}")
            column_indices[header.strip().lower()] = idx
    print("Column Indices:", column_indices)  # Debug: Print complete dictionary
    return column_indices

def map_headers_to_required_fields(column_indices, required_fields):
    """Map detected headers to required fields using flexible matching."""
    header_mapping = {}
    for field, possible_names in required_fields.items():
        for name in possible_names:
            for header in column_indices:
                if name.lower() in header.lower():
                    header_mapping[field] = column_indices[header]
                    break
            else:
                continue
            break
        else:
            print(f"Warning: Required field '{field}' not found in the input file. Possible names: {possible_names}")
            header_mapping[field] = None
    return header_mapping

def populate_instrument_template(new_sheet, row_data, header_mapping):
    """Populate the Instrument template with data from the row.
       For instruments, the control signal goes into cell D14."""
    tag_val = row_data[header_mapping["tag"]]
    if tag_val is not None:
        tag_val = tag_val.splitlines()[0].strip()
    new_sheet["A11"] = format_value(tag_val)
    new_sheet["C11"] = format_value(row_data[header_mapping["manufacturer"]])
    new_sheet["E11"] = format_value(row_data[header_mapping["model"]])
    new_sheet["A14"] = format_value(row_data[header_mapping["process connection"]])
    new_sheet["B14"] = format_value(row_data[header_mapping["immersion length"]])
    new_sheet["D14"] = format_value(row_data[header_mapping["control signal"]])  # Instrument: D14
    new_sheet["F14"] = format_value(row_data[header_mapping["min range"]])
    new_sheet["G14"] = format_value(row_data[header_mapping["max range"]])
    new_sheet["H14"] = format_value(row_data[header_mapping["unit"]])
    new_sheet["G11"] = format_value(row_data[header_mapping["order code"]])
    new_sheet["I14"] = "N/A"

def populate_valve_template(new_sheet, row_data, header_mapping):
    """Populate the Valve template with data from the row.
       For valves, the control signal goes into cell D15."""
    tag_val = row_data[header_mapping["tag"]]
    if tag_val is not None:
        tag_val = tag_val.splitlines()[0].strip()
    new_sheet["A11"] = format_value(tag_val)
    new_sheet["C11"] = format_value(row_data[header_mapping["valve make / model number"]])
    new_sheet["F11"] = format_value(row_data[header_mapping["actuator make / model number"]])
    new_sheet["A15"] = format_value(row_data[header_mapping["process connection"]])
    new_sheet["B15"] = format_value(row_data[header_mapping["line size"]])
    new_sheet["D15"] = format_value(row_data[header_mapping["control signal"]])  # Valve: D15
    new_sheet["F15"] = format_value(row_data[header_mapping["dial setting"]])
    new_sheet["I15"] = format_value(row_data[header_mapping["flow rate"]])

def apply_formatting(new_sheet):
    """Apply alignment and font settings to all populated cells."""
    font = Font(name="Calibri", size=FONT_SIZE, bold=False)
    for cell_ref in [
        "A5", "E5", "A7", "E7", "I5", "I7", "A11", "C11", "E11",
        "A14", "B14", "D14", "F14", "G14", "H14", "G11", "I14",
        "F11", "A15", "B15", "F15", "I15", "D15"
    ]:
        cell = new_sheet[cell_ref]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = font

def generate_rv_forms(input_file, output_file, project, client, reference_document,
                      document_revision, start_row, template_type, progress_var, log_file):
    try:
        print(f"Loading input file: {input_file}")
        wb = openpyxl.load_workbook(input_file)

        if template_type == "Instrument":
            template_sheet_name = get_sheet_by_partial_name(wb, "RV Instrument  SUB-TF-01")
            required_fields = {
                "tag": ["Tag", "Instrument Tag", "BMS Tag"],
                "manufacturer": ["Manufacturer", "Instrument Manufacturer"],
                "model": ["model", "instrument model", "model number"],
                "process connection": ["process connection", "connection"],
                "immersion length": ["immersion length", "Immersion Length (mm)"],
                "control signal": ["control signal", "actuator control signal"],
                "min range": ["min range", "range min", "minimum range"],
                "max range": ["max range", "range max", "maximum range"],
                "unit": ["unit", "units"],
                "order code": ["order code", "code"],
            }
        else:
            template_sheet_name = get_sheet_by_partial_name(wb, "RV Valve SUB-TF-02")
            required_fields = {
                "tag": ["BMS Tag", "Instrument Tag"],
                "valve make / model number": ["valve make", "valve model", "valve make / model number"],
                "actuator make / model number": ["actuator make", "actuator model", "actuator make / model number"],
                "process connection": ["process connection", "connection"],
                "line size": ["line size", "line size (mm)", "Line Size", "Valve Size", "Valve Size[mm]"],
                "control signal": ["control signal", "actuator control signal", "signal type", "Valve Signal", "Actuator Control signal"],
                "dial setting": ["dial setting", "setting", "dial"],
                "flow rate": ["flow rate", "rate", "flow", "Flow Rate"],
            }

        template_sheet = wb[template_sheet_name]
        current_date = datetime.now().strftime("%d %b %Y").upper()

        with open(log_file, "a") as log:
            log.write(f"\n{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Starting RV form generation for project: {project}\n")

            sheet1 = wb[wb.sheetnames[0]]

            try:
                header_row = detect_header_row(sheet1)
                log.write(f"Detected header row: {header_row}\n")
            except ValueError as e:
                if start_row:
                    header_row = int(start_row)
                    log.write(f"Auto-detect failed. Using fallback header row: {header_row}\n")
                else:
                    raise e

            column_indices = get_column_indices(sheet1, header_row)
            header_mapping = map_headers_to_required_fields(column_indices, required_fields)

            print("Header Mapping:", header_mapping)  # Debug output

            check_field = "tag"
            if header_mapping.get(check_field) is None:
                raise ValueError(f"Required field '{check_field}' not found. Possible names: {required_fields[check_field]}")

            total_rows = sum(1 for _ in sheet1.iter_rows(min_row=header_row + 1, values_only=True)
                             if _[header_mapping[check_field]])
            progress_step = 100 / total_rows if total_rows > 0 else 100
            progress = 0
            progress_var.set(0)

            instrument_no_index = get_instrument_no_index(column_indices)
            if instrument_no_index is None:
                raise ValueError("Unique identifier column not found (expected 'No.', 'No', 'Item No.' or 'Item No').")

            processed_ids = set()
            rv_counter = 1

            for row in sheet1.iter_rows(min_row=header_row + 1, values_only=True):
                tag_val = row[header_mapping[check_field]]
                if tag_val is None:
                    print("Skipping row because the tag field is missing.")
                    continue

                instrument_number = row[instrument_no_index]
                if not instrument_number:
                    print("Skipping row due to empty instrument number (likely an extra tag row).")
                    continue
                if instrument_number in processed_ids:
                    print(f"Skipping duplicate row for instrument number: {instrument_number}")
                    continue
                processed_ids.add(instrument_number)

                try:
                    new_sheet = wb.copy_worksheet(template_sheet)
                    rv_form_name = f"RV{rv_counter:02d}"  # Consistent two-digit formatting
                    new_sheet.title = rv_form_name

                    if template_type == "Instrument":
                        populate_instrument_template(new_sheet, row, header_mapping)
                    else:
                        populate_valve_template(new_sheet, row, header_mapping)

                    new_sheet["A5"] = project
                    new_sheet["E5"] = client
                    new_sheet["A7"] = reference_document
                    new_sheet["E7"] = document_revision
                    new_sheet["I5"] = current_date
                    new_sheet["I7"] = rv_form_name

                    apply_formatting(new_sheet)
                    
                    # --- Insert the template logo into cell C1 so it spans columns C to F ---
                    try:
                        templatelogo_path = resource_path(os.path.join("resources", "templatelogo.png"))
                        if os.path.exists(templatelogo_path):
                            tmpl_img = XLImage(templatelogo_path)
                            # Calculate dimensions:
                            # 1.67 cm in points: (1.67/2.54)*72 ≈ 47.3, scaled by 1.35 gives ≈ 63.9 points (use 64)
                            # 7.22 cm in points: (7.22/2.54)*72 ≈ 204.7, scaled by 1.35 gives ≈ 276.3 points (use 276)
                            tmpl_img.height = 64
                            tmpl_img.width = 276
                            tmpl_img.rotation = 0
                            tmpl_img.anchor = "C1"  # Anchor at C1 (so it spans from C to F as desired)
                            new_sheet.add_image(tmpl_img)
                        else:
                            print("Template logo file not found at:", templatelogo_path)
                    except Exception as img_err:
                        print("Error inserting template logo:", img_err)
                    # --- End of template logo insertion ---

                    log.write(f"Processed: {rv_form_name} (Instrument Number: {instrument_number})\n")
                    rv_counter += 1

                except Exception as e:
                    log.write(f"Error processing a row: {str(e)}\n")

                progress += progress_step
                progress_var.set(progress)

            template_sheet.sheet_state = 'hidden'

            try:
                print(f"Saving output file: {output_file}")
                wb.save(output_file)
                print("File saved successfully.")
            except Exception as e:
                raise Exception(f"Failed to save the file: {str(e)}")

            progress_var.set(100)
            log.write("RV form generation completed successfully.\n")
            messagebox.showinfo("Success", f"RV forms generated successfully! Output saved as {output_file}")

    except Exception as e:
        with open(log_file, "a") as log:
            log.write(f"Error: {str(e)}\n")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# --- GUI Implementation ---
def main():
    def browse_input_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")])
        input_file_var.set(file_path)

    def browse_output_location():
        folder_path = filedialog.askdirectory()
        output_folder_var.set(folder_path)

    def generate_forms():
        input_file = input_file_var.get()
        project = project_var.get()
        client = client_var.get()
        reference_document = reference_var.get()
        document_revision = revision_var.get()
        start_row = start_row_var.get().strip()  # Optional fallback header row.
        output_folder = output_folder_var.get()
        template_type = template_type_var.get()

        if not input_file or not project or not client or not reference_document or not document_revision or not output_folder or not template_type:
            messagebox.showerror("Input Error", "Please fill in all fields and select files.")
            return

        file_extension = os.path.splitext(input_file)[-1]
        output_file = os.path.join(output_folder, f"{project}-RVs{file_extension}")
        log_file = os.path.join(output_folder, "rv_generator_log.txt")

        try:
            generate_rv_forms(input_file, output_file, project, client, reference_document,
                              document_revision, start_row, template_type, progress_var, log_file)
        except Exception as e:
            with open(log_file, "a") as log:
                log.write(f"Error: {str(e)}\n")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    root = tk.Tk()
    root.title("RV Form Generator")
    root.configure(bg="#ffffff")
    root.geometry("600x500")

    try:
        icon_path = os.path.join("resources", "subnetlogo.ico")
        root.iconbitmap(icon_path)
    except Exception as e:
        print("Window icon not set:", e)

    input_file_var = tk.StringVar()
    output_folder_var = tk.StringVar()
    project_var = tk.StringVar()
    client_var = tk.StringVar()
    reference_var = tk.StringVar()
    revision_var = tk.StringVar()
    start_row_var = tk.StringVar(value="")  # Optional fallback.
    template_type_var = tk.StringVar(value="Instrument")
    progress_var = tk.DoubleVar()

    try:
        logo_path = os.path.join("resources", "subnetlogo.png")
        logo_image = PhotoImage(file=logo_path)
        logo_label = tk.Label(root, image=logo_image, bg="#ffffff")
        logo_label.grid(row=0, column=0, columnspan=3, pady=(10, 20))
    except Exception as e:
        print("Logo image not found:", e)

    pad_options = {'padx': 10, 'pady': 5}
    tk.Label(root, text="Input File:", bg="#ffffff", font=("Arial", 10)).grid(row=1, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=input_file_var, width=50).grid(row=1, column=1, **pad_options)
    tk.Button(root, text="Browse", command=browse_input_file, bg="#4CAF50", fg="#ffffff", font=("Arial", 10)).grid(row=1, column=2, **pad_options)

    tk.Label(root, text="Output Location:", bg="#ffffff", font=("Arial", 10)).grid(row=2, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=output_folder_var, width=50).grid(row=2, column=1, **pad_options)
    tk.Button(root, text="Browse", command=browse_output_location, bg="#4CAF50", fg="#ffffff", font=("Arial", 10)).grid(row=2, column=2, **pad_options)

    tk.Label(root, text="Template Type:", bg="#ffffff", font=("Arial", 10)).grid(row=3, column=0, sticky="e", **pad_options)
    template_dropdown = ttk.Combobox(root, textvariable=template_type_var, values=["Instrument", "Valve"], state="readonly", width=47)
    template_dropdown.grid(row=3, column=1, columnspan=2, **pad_options)

    tk.Label(root, text="Project Name:", bg="#ffffff", font=("Arial", 10)).grid(row=4, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=project_var, width=50).grid(row=4, column=1, columnspan=2, **pad_options)

    tk.Label(root, text="Client Name:", bg="#ffffff", font=("Arial", 10)).grid(row=5, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=client_var, width=50).grid(row=5, column=1, columnspan=2, **pad_options)

    tk.Label(root, text="Reference Document:", bg="#ffffff", font=("Arial", 10)).grid(row=6, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=reference_var, width=50).grid(row=6, column=1, columnspan=2, **pad_options)

    tk.Label(root, text="Document Revision:", bg="#ffffff", font=("Arial", 10)).grid(row=7, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=revision_var, width=50).grid(row=7, column=1, columnspan=2, **pad_options)

    tk.Label(root, text="Header Row Fallback (optional):", bg="#ffffff", font=("Arial", 10)).grid(row=8, column=0, sticky="e", **pad_options)
    tk.Entry(root, textvariable=start_row_var, width=50).grid(row=8, column=1, columnspan=2, **pad_options)

    tk.Button(root, text="Generate RV Forms", command=generate_forms, bg="#4CAF50", fg="#ffffff", font=("Arial", 12, "bold")).grid(row=9, column=0, columnspan=3, pady=15)

    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
    progress_bar.grid(row=10, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

    tk.Label(root, text="© 2025 Subnet Ltd. All rights reserved.", bg="#ffffff", font=("Arial", 9)).grid(row=11, column=0, columnspan=3, pady=(10, 0))

    root.mainloop()

if __name__ == "__main__":
    main()



